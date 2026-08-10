#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""kb_data.json から Obsidian 風ナレッジベースの Excel ブックを生成する。

使い方:
    python3 build_kb.py <kb_data.json> <output.xlsx>

kb_data.json のスキーマ:
{
  "version": 1,
  "updated": "YYYY-MM-DD",
  "notes": [
    {
      "id":      "M-20260810-01",        # 一意ID (M-=メール, C-=Claude, H-=ハブ, N-=メモ)
      "date":    "YYYY-MM-DD",
      "type":    "メール" | "Claude" | "ハブ" | "メモ",
      "source":  "差出人やセッション名など",
      "title":   "タイトル",
      "summary": "1〜2行の要約",
      "tags":    ["転職", "スカウト"],
      "links":   [{"to": "H-転職活動", "rel": "トピック"}],
      "url":     "https://... (Gmailスレッドやセッションへのリンク、省略可)",
      "ref":     "Gmail thread id など重複判定用の外部ID (省略可)"
    }
  ]
}

生成されるシート:
  ダッシュボード / ノート / リンク / タグ
バックリンク一覧は Python 側で導出し、集計値 (件数など) は数式で計算する。
"""
import json
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
TYPE_FILL = {
    "ハブ": PatternFill("solid", fgColor="FEF3C7"),
    "メール": PatternFill("solid", fgColor="DBEAFE"),
    "Claude": PatternFill("solid", fgColor="EDE9FE"),
    "メモ": PatternFill("solid", fgColor="DCFCE7"),
}
LINK_FONT = Font(name=FONT_NAME, size=10, color="1155CC", underline="single")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)

    # ハブを先頭に、それ以外は新しい日付順で表示する
    hubs = sorted((n for n in data["notes"] if n["type"] == "ハブ"),
                  key=lambda n: n["id"])
    rest = sorted(
        (n for n in data["notes"] if n["type"] != "ハブ"),
        key=lambda n: (n.get("date", ""), n["id"]),
        reverse=True,
    )
    notes = hubs + rest

    by_id = {n["id"]: n for n in notes}
    backlinks = {}
    edges = []
    for n in notes:
        for link in n.get("links", []):
            to = link["to"]
            if to not in by_id:
                print(f"警告: {n['id']} のリンク先 {to} が存在しません", file=sys.stderr)
                continue
            edges.append((n["id"], link.get("rel", "関連"), to))
            backlinks.setdefault(to, []).append(n["id"])

    row_of = {n["id"]: i + 2 for i, n in enumerate(notes)}

    wb = Workbook()

    # ---- ノート ----
    ws = wb.active
    ws.title = "ノート"
    headers = [
        "ID", "日付", "種別", "出所", "タイトル", "要約",
        "タグ", "リンク先", "バックリンク", "被リンク数", "URL",
    ]
    ws.append(headers)
    for n in notes:
        r = row_of[n["id"]]
        links_txt = ", ".join(
            f"{l['to']} ({l.get('rel', '関連')})" for l in n.get("links", [])
        )
        back_txt = ", ".join(backlinks.get(n["id"], []))
        ws.cell(row=r, column=1, value=n["id"])
        ws.cell(row=r, column=2, value=n.get("date", ""))
        ws.cell(row=r, column=3, value=n["type"])
        ws.cell(row=r, column=4, value=n.get("source", ""))
        ws.cell(row=r, column=5, value=n["title"])
        ws.cell(row=r, column=6, value=n.get("summary", ""))
        ws.cell(row=r, column=7, value=", ".join(n.get("tags", [])))
        ws.cell(row=r, column=8, value=links_txt)
        ws.cell(row=r, column=9, value=back_txt)
        ws.cell(row=r, column=10, value=f"=COUNTIF(リンク!$D:$D,A{r})")
        if n.get("url"):
            c = ws.cell(row=r, column=11, value="開く")
            c.hyperlink = n["url"]
            c.font = LINK_FONT
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            if cell.font is None or cell.font.name != FONT_NAME or col != 11:
                if col != 11:
                    cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (6, 8, 9)))
        ws.cell(row=r, column=3).fill = TYPE_FILL.get(n["type"], PatternFill())
    style_header(ws, len(headers))
    set_widths(ws, [15, 11, 8, 24, 40, 50, 20, 30, 26, 10, 8])
    ws.auto_filter.ref = f"A1:K{len(notes) + 1}"

    # ---- リンク ----
    ws = wb.create_sheet("リンク")
    ws.append(["From ID", "From タイトル", "関係", "To ID", "To タイトル"])
    for i, (frm, rel, to) in enumerate(edges, start=2):
        c = ws.cell(row=i, column=1, value=frm)
        c.hyperlink = f"#ノート!A{row_of[frm]}"
        c.font = LINK_FONT
        ws.cell(row=i, column=2,
                value=f"=IFERROR(INDEX(ノート!$E:$E,MATCH(A{i},ノート!$A:$A,0)),\"\")")
        ws.cell(row=i, column=3, value=rel)
        c = ws.cell(row=i, column=4, value=to)
        c.hyperlink = f"#ノート!A{row_of[to]}"
        c.font = LINK_FONT
        ws.cell(row=i, column=5,
                value=f"=IFERROR(INDEX(ノート!$E:$E,MATCH(D{i},ノート!$A:$A,0)),\"\")")
        for col in (2, 3, 5):
            ws.cell(row=i, column=col).font = Font(name=FONT_NAME, size=10)
    style_header(ws, 5)
    set_widths(ws, [15, 42, 16, 15, 42])
    ws.auto_filter.ref = f"A1:E{len(edges) + 1}"

    # ---- タグ ----
    tag_notes = OrderedDict()
    for n in notes:
        for t in n.get("tags", []):
            tag_notes.setdefault(t, []).append(n["id"])
    tags_sorted = sorted(tag_notes.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    ws = wb.create_sheet("タグ")
    ws.append(["タグ", "件数", "ノートID"])
    for i, (tag, ids) in enumerate(tags_sorted, start=2):
        ws.cell(row=i, column=1, value=tag).font = Font(name=FONT_NAME, size=10, bold=True)
        # タグ列はカンマ区切り文字列のためワイルドカード一致で数える
        # (タグ名同士が部分文字列にならない語彙運用を前提とする)
        ws.cell(row=i, column=2, value=f'=COUNTIF(ノート!$G:$G,"*"&A{i}&"*")')
        c = ws.cell(row=i, column=3, value=", ".join(ids))
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws, 3)
    set_widths(ws, [18, 8, 90])

    # ---- ダッシュボード ----
    ws = wb.create_sheet("ダッシュボード", 0)
    title = ws.cell(row=1, column=1, value="ナレッジベース ダッシュボード")
    title.font = Font(name=FONT_NAME, size=16, bold=True)
    ws.cell(row=2, column=1, value=f"最終更新: {data.get('updated', '')}").font = Font(
        name=FONT_NAME, size=10, italic=True, color="666666")

    rows = [
        ("ノート総数", "=COUNTA(ノート!$A:$A)-1"),
        ("メール", '=COUNTIF(ノート!$C:$C,"メール")'),
        ("Claudeセッション", '=COUNTIF(ノート!$C:$C,"Claude")'),
        ("ハブ (MOC)", '=COUNTIF(ノート!$C:$C,"ハブ")'),
        ("メモ", '=COUNTIF(ノート!$C:$C,"メモ")'),
        ("リンク総数", "=COUNTA(リンク!$A:$A)-1"),
        ("タグ数", "=COUNTA(タグ!$A:$A)-1"),
        ("最多被リンク数", "=MAX(ノート!$J:$J)"),
        ("最多被リンクノート",
         "=INDEX(ノート!$E:$E,MATCH(MAX(ノート!$J:$J),ノート!$J:$J,0))"),
    ]
    for i, (label, formula) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=True)
        ws.cell(row=i, column=2, value=formula).font = Font(name=FONT_NAME, size=10)

    legend = [
        "使い方:",
        "・「ノート」= Obsidianのノート。1行が1ノートで、メール/Claudeセッション/ハブが混在します。",
        "・「リンク先」「バックリンク」= Obsidianの [[リンク]] とバックリンクに相当します。",
        "・「リンク」シートがグラフの辺 (エッジ) 一覧です。IDクリックでノート行へジャンプします。",
        "・「タグ」シートが #タグ の索引です。",
        "・ハブ (MOC) はトピックのまとめノートです。新しいノートはまずハブに紐付けます。",
        "・このファイルは毎日のルーティンで自動再生成されます。手で編集する場合は kb_data.json 側へ。",
    ]
    for i, line in enumerate(legend, start=15):
        ws.cell(row=i, column=1, value=line).font = Font(name=FONT_NAME, size=9, color="444444")
    set_widths(ws, [26, 46])

    wb.save(sys.argv[2])
    print(f"OK: {len(notes)} ノート / {len(edges)} リンク / {len(tags_sorted)} タグ -> {sys.argv[2]}")


if __name__ == "__main__":
    main()
